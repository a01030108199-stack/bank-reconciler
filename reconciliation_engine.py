import pandas as pd
import numpy as np
from datetime import datetime
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def reconcile_statements(df_bank_raw, df_ledger_raw):
    # Copy data to avoid side effects
    df_bank = df_bank_raw.copy()
    df_ledger = df_ledger_raw.copy()
    
    # Ensure columns exist, fill NaNs
    for col in ['المدين (مسحوبات)', 'الدائن (إيداعات)']:
        if col in df_bank.columns:
            df_bank[col] = df_bank[col].fillna(0.0).astype(float)
    for col in ['المدين', 'الدائن']:
        if col in df_ledger.columns:
            df_ledger[col] = df_ledger[col].fillna(0.0).astype(float)
            
    # Parse dates
    df_bank['date_parsed'] = pd.to_datetime(df_bank['التاريخ'])
    df_ledger['date_parsed'] = pd.to_datetime(df_ledger['التاريخ'])
    
    # Calculate Net Amounts
    # Bank: Deposits (Credit) - Withdrawals (Debit)
    df_bank['net_amount'] = df_bank['الدائن (إيداعات)'] - df_bank['المدين (مسحوبات)']
    # Ledger: Debit - Credit
    df_ledger['net_amount'] = df_ledger['المدين'] - df_ledger['الدائن']
    
    # Initialize match status
    df_bank['matched'] = False
    df_bank['match_id'] = ""
    df_bank['notes'] = ""
    
    df_ledger['matched'] = False
    df_ledger['match_id'] = ""
    df_ledger['notes'] = ""
    
    match_counter = 1
    
    # Matching pass 1: Match by Amount and Reference
    # Reference in Bank is 'المرجع', Doc Num in Ledger is 'رقم القيد/المستند'
    for b_idx, b_row in df_bank.iterrows():
        if b_row['matched']:
            continue
        ref = str(b_row['المرجع']).strip()
        if not ref or ref.lower() == 'nan':
            continue
            
        # Find matching ledger row by Net Amount and Reference/Doc Num
        match_l = df_ledger[
            (~df_ledger['matched']) & 
            (df_ledger['net_amount'] == b_row['net_amount']) & 
            (df_ledger['رقم القيد/المستند'].astype(str).str.strip() == ref)
        ]
        
        if not match_l.empty:
            l_idx = match_l.index[0]
            match_id = f"MCH-{match_counter:04d}"
            match_counter += 1
            
            df_bank.at[b_idx, 'matched'] = True
            df_bank.at[b_idx, 'match_id'] = match_id
            df_bank.at[b_idx, 'notes'] = "مطابقة بالمرجع والقيمة"
            
            df_ledger.at[l_idx, 'matched'] = True
            df_ledger.at[l_idx, 'match_id'] = match_id
            df_ledger.at[l_idx, 'notes'] = "مطابقة بالمرجع والقيمة"
            
    # Matching pass 2: Match by Amount and Date (within +/- 4 days window)
    for b_idx, b_row in df_bank.iterrows():
        if b_row['matched']:
            continue
            
        b_date = b_row['date_parsed']
        b_amount = b_row['net_amount']
        
        # Filter ledger records by same net amount and date difference
        candidates = df_ledger[
            (~df_ledger['matched']) & 
            (df_ledger['net_amount'] == b_amount)
        ].copy()
        
        if not candidates.empty:
            candidates['date_diff'] = (candidates['date_parsed'] - b_date).abs().dt.days
            candidates = candidates.sort_values(by='date_diff')
            
            # If the closest transaction is within 4 days, match it
            if candidates.iloc[0]['date_diff'] <= 4:
                l_idx = candidates.index[0]
                match_id = f"MCH-{match_counter:04d}"
                match_counter += 1
                
                df_bank.at[b_idx, 'matched'] = True
                df_bank.at[b_idx, 'match_id'] = match_id
                df_bank.at[b_idx, 'notes'] = f"مطابقة تلقائية بفارق تاريخ {candidates.iloc[0]['date_diff']} يوم"
                
                df_ledger.at[l_idx, 'matched'] = True
                df_ledger.at[l_idx, 'match_id'] = match_id
                df_ledger.at[l_idx, 'notes'] = f"مطابقة تلقائية بفارق تاريخ {candidates.iloc[0]['date_diff']} يوم"
                
    # Create Reconciled DataFrames
    df_matched_bank = df_bank[df_bank['matched']].copy()
    df_unmatched_bank = df_bank[~df_bank['matched']].copy()
    
    df_matched_ledger = df_ledger[df_ledger['matched']].copy()
    df_unmatched_ledger = df_ledger[~df_ledger['matched']].copy()
    
    return df_matched_bank, df_unmatched_bank, df_matched_ledger, df_unmatched_ledger

def generate_reconciliation_summary(df_bank, df_ledger, df_unmatched_bank, df_unmatched_ledger):
    # Ending Balances
    bank_ending_balance = float(df_bank.iloc[-1]['الرصيد']) if not df_bank.empty else 0.0
    ledger_ending_balance = float(df_ledger.iloc[-1]['الرصيد']) if not df_ledger.empty else 0.0
    
    # 1. Outstanding Checks (شيكات صادرة لم تقدم للصرف)
    # Ledger outflows (Credit > 0 or net_amount < 0) that are unmatched in bank
    outstanding_checks_df = df_unmatched_ledger[df_unmatched_ledger['net_amount'] < 0]
    total_outstanding_checks = abs(outstanding_checks_df['net_amount'].sum())
    
    # 2. Deposits in Transit (إيداعات بالطريق)
    # Ledger inflows (Debit > 0 or net_amount > 0) that are unmatched in bank
    deposits_in_transit_df = df_unmatched_ledger[df_unmatched_ledger['net_amount'] > 0]
    total_deposits_in_transit = deposits_in_transit_df['net_amount'].sum()
    
    # 3. Direct Credits in Bank (إيداعات مباشرة بالبنك لم تسجل بالدفاتر)
    # Bank inflows (Credit/Deposits > 0 or net_amount > 0) unmatched in ledger
    direct_credits_df = df_unmatched_bank[df_unmatched_bank['net_amount'] > 0]
    total_direct_credits = direct_credits_df['net_amount'].sum()
    
    # 4. Bank Charges / Interest (عمولات ومصاريف بنكية لم تسجل بالدفاتر)
    # Bank outflows (Debit/Withdrawals > 0 or net_amount < 0) unmatched in ledger
    bank_charges_df = df_unmatched_bank[df_unmatched_bank['net_amount'] < 0]
    total_bank_charges = abs(bank_charges_df['net_amount'].sum())
    
    # Calculations
    adjusted_bank_balance = bank_ending_balance + total_deposits_in_transit - total_outstanding_checks
    adjusted_ledger_balance = ledger_ending_balance + total_direct_credits - total_bank_charges
    unreconciled_difference = adjusted_bank_balance - adjusted_ledger_balance
    
    summary = {
        'bank_ending_balance': bank_ending_balance,
        'ledger_ending_balance': ledger_ending_balance,
        'total_outstanding_checks': total_outstanding_checks,
        'outstanding_checks_df': outstanding_checks_df,
        'total_deposits_in_transit': total_deposits_in_transit,
        'deposits_in_transit_df': deposits_in_transit_df,
        'total_direct_credits': total_direct_credits,
        'direct_credits_df': direct_credits_df,
        'total_bank_charges': total_bank_charges,
        'bank_charges_df': bank_charges_df,
        'adjusted_bank_balance': adjusted_bank_balance,
        'adjusted_ledger_balance': adjusted_ledger_balance,
        'unreconciled_difference': unreconciled_difference
    }
    
    return summary

def export_reconciliation_excel(summary, df_matched_bank, df_unmatched_bank, df_matched_ledger, df_unmatched_ledger):
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Summary Statement
        summary_rows = [
            {"البند": "رصيد الحساب الجاري حسب كشف حساب البنك", "القيمة (جنيه مصري)": summary['bank_ending_balance']},
            {"البند": "(+) يضاف: الإيداعات بالطريق (إيداعات غير ظاهرة بالبنك)", "القيمة (جنيه مصري)": summary['total_deposits_in_transit']},
            {"البند": "(-) يخصم: شيكات صادرة معلقة (لم تقدم للصرف)", "القيمة (جنيه مصري)": -summary['total_outstanding_checks']},
            {"البند": "رصيد البنك المعدل المطبق", "القيمة (جنيه مصري)": summary['adjusted_bank_balance']},
            {"البند": "", "القيمة (جنيه مصري)": None},
            {"البند": "رصيد النقدية بالبنك حسب دفاتر الأستاذ العام", "القيمة (جنيه مصري)": summary['ledger_ending_balance']},
            {"البند": "(+) يضاف: المقبوضات والتحويلات المباشرة بالبنك", "القيمة (جنيه مصري)": summary['total_direct_credits']},
            {"البند": "(-) يخصم: العمولات والمصاريف البنكية غير المسجلة", "القيمة (جنيه مصري)": -summary['total_bank_charges']},
            {"البند": "رصيد الدفاتر المعدل المطبق", "القيمة (جنيه مصري)": summary['adjusted_ledger_balance']},
            {"البند": "", "القيمة (جنيه مصري)": None},
            {"البند": "الفروقات غير المسواة (يجب أن تساوي 0)", "القيمة (جنيه مصري)": summary['unreconciled_difference']}
        ]
        
        df_sum_sheet = pd.DataFrame(summary_rows)
        df_sum_sheet.to_excel(writer, sheet_name="مذكرة التسوية", index=False)
        
        # Sheet 2: Unmatched Bank (Exceptions Bank)
        if not df_unmatched_bank.empty:
            df_unmatched_bank[['التاريخ', 'البيان', 'المرجع', 'المدين (مسحوبات)', 'الدائن (إيداعات)', 'الرصيد']].to_excel(
                writer, sheet_name="معلقات البنك", index=False
            )
        else:
            pd.DataFrame([{"رسالة": "لا توجد معلقات في كشف البنك"}]).to_excel(writer, sheet_name="معلقات البنك", index=False)
            
        # Sheet 3: Unmatched Ledger (Exceptions Ledger)
        if not df_unmatched_ledger.empty:
            df_unmatched_ledger[['التاريخ', 'رقم القيد/المستند', 'الشرح', 'المدين', 'الدائن', 'الرصيد']].to_excel(
                writer, sheet_name="معلقات الدفاتر", index=False
            )
        else:
            pd.DataFrame([{"رسالة": "لا توجد معلقات في الدفاتر"}]).to_excel(writer, sheet_name="معلقات الدفاتر", index=False)
            
        # Sheet 4: Matched Transactions
        df_matched_bank_clean = df_matched_bank[['التاريخ', 'البيان', 'المرجع', 'net_amount', 'match_id', 'notes']].copy()
        df_matched_bank_clean.columns = ['التاريخ', 'البيان', 'المرجع', 'المبلغ الصافي', 'كود المطابقة', 'ملاحظات']
        df_matched_bank_clean.to_excel(writer, sheet_name="العمليات المطابقة", index=False)
        
        # Apply openpyxl styling
        workbook = writer.book
        
        # Colors
        fill_header = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid") # Dark Purple/Violet
        fill_sub_header = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid") # Light Purple
        fill_success = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Soft Green
        fill_alert = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
        fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Zebra Light Gray
        
        # Fonts
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=11, bold=True)
        font_normal = Font(name="Arial", size=11)
        font_green = Font(name="Arial", size=11, color="065F46", bold=True)
        font_red = Font(name="Arial", size=11, color="991B1B", bold=True)
        
        # Borders
        thin_side = Side(border_style="thin", color="D1D5DB")
        double_side = Side(border_style="double", color="374151")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thin_side, bottom=double_side)
        
        # Alignments
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Format Sheet 1: مذكرة التسوية
        ws_sum = workbook["مذكرة التسوية"]
        ws_sum.views.sheetView[0].showGridLines = True
        ws_sum.column_dimensions['A'].width = 55
        ws_sum.column_dimensions['B'].width = 22
        
        # Header Row
        ws_sum.row_dimensions[1].height = 25
        for col_idx in [1, 2]:
            cell = ws_sum.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center if col_idx == 2 else align_right
            
        # Data Rows (Rows 2 to 12)
        for r_idx in range(2, 13):
            ws_sum.row_dimensions[r_idx].height = 24
            cell_a = ws_sum.cell(row=r_idx, column=1)
            cell_b = ws_sum.cell(row=r_idx, column=2)
            
            # Text and alignment
            cell_a.font = font_normal
            cell_a.alignment = align_right
            cell_b.alignment = align_left
            cell_b.number_format = '#,##0.00'
            
            # Alternating Zebra
            if r_idx % 2 == 1:
                cell_a.fill = fill_zebra
                cell_b.fill = fill_zebra
                
            val_a = str(cell_a.value)
            
            # Styling specific accounting rows
            if "(+)" in val_a:
                cell_a.font = Font(name="Arial", size=11, color="10B981")
                cell_b.font = Font(name="Arial", size=11, color="10B981")
            elif "(-)" in val_a:
                cell_a.font = Font(name="Arial", size=11, color="EF4444")
                cell_b.font = Font(name="Arial", size=11, color="EF4444")
            elif "المعدل" in val_a or "رصيد" in val_a:
                cell_a.font = font_bold
                cell_b.font = font_bold
                cell_a.fill = fill_sub_header
                cell_b.fill = fill_sub_header
                cell_a.border = border_total
                cell_b.border = border_total
            elif "الفروقات" in val_a:
                cell_a.font = font_bold
                cell_b.font = font_bold
                diff = summary['unreconciled_difference']
                if abs(diff) < 0.01:
                    cell_a.fill = fill_success
                    cell_b.fill = fill_success
                    cell_a.font = font_green
                    cell_b.font = font_green
                else:
                    cell_a.fill = fill_alert
                    cell_b.fill = fill_alert
                    cell_a.font = font_red
                    cell_b.font = font_red
                    
        # Format Sheet 2, 3, 4: Data Tables
        for sheet_name in ["معلقات البنك", "معلقات الدفاتر", "العمليات المطابقة"]:
            ws = workbook[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            # Header Styling
            ws.row_dimensions[1].height = 25
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align_center
                
            # Row Styling and Auto-width
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                is_even = (row_idx % 2 == 0)
                
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_normal
                    cell.border = border_all
                    
                    if is_even:
                        cell.fill = fill_zebra
                        
                    # Alignment and number formats
                    val = cell.value
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = align_left
                    elif isinstance(val, datetime) or (isinstance(val, str) and '-' in val and len(val) == 10):
                        cell.alignment = align_center
                    elif isinstance(val, str) and (val.startswith('REF-') or val.startswith('JV-') or val.startswith('MCH-') or val.startswith('CHG-') or val.startswith('INT-')):
                        cell.alignment = align_center
                        if val.startswith('MCH-'):
                            cell.font = font_bold
                    else:
                        cell.alignment = align_right
                        
            # Auto-fit Column Widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
    excel_data = output.getvalue()
    return excel_data


